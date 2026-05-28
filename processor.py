"""
PO Cost Changes — pandas translation of the Power Query Section1.m pipeline.

Original M lives in ../docs/Section1.m. This module mirrors it step-for-step
so changes can be diffed against the source. If you change a step here,
update the step number reference too.
"""
from __future__ import annotations

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from mapping import get_mapping
from teams import get_teams, rename_team_vendor
from vendors import get_open_vendors, offset_category

# QBO Company → display label.
# Used for the Company column value in output files, the per-company file
# names, and the UI summary table. Edit here when company labels change.
DISPLAY_NAMES: dict[str, str] = {
    "Y&S Tickets":         "Y&S",
    "Damona & Crew":       "Damona",
    "The Ticket Guy LLC":  "The Ticket Guy",
    "YourTickets":         "YourTickets",
    "YS Asher Tickets":    "YSA",
    "YS Chase Tickets":    "Chase (Jacks)",
    "YS Katz Tickets":     "Katz",
    "YS Levine Tickets":   "Levine",
    "YS Levovitz Tickets": "Levovitz",
    "YS Needle Tickets":   "Needle",
    "YS TL Tickets":       "TL",
    "YSKG Tickets":        "YSKG",
    "YSM Tickets":         "Grossman",
    "YSP Tickets":         "Pollak",
    "YSS Tickets":         "Sternbuch",
    "YSW Tickets":         "YSW (Waxler)",
}

# Sheet/file order for the combined workbook and UI grid. Matches the
# Purchase Details processor's ordering. Any display label not listed here
# falls to the end (alphabetical), and "YourTickets" always goes last.
DISPLAY_ORDER: list[str] = [
    "Y&S",
    "Grossman",
    "Sternbuch",
    "Pollak",
    "Levine",
    "Levovitz",
    "YSKG",
    "The Ticket Guy",
    "Chase (Jacks)",
    "YSA",
    "Katz",
    "Needle",      # not in reference; slotted with the other affiliates
    "TL",
    "YSW (Waxler)",
    "Damona",
    "YourTickets",  # always last
]


def display_name(qbo_company: str) -> str:
    """Return the display label for a QBO company, or the QBO name itself
    if there's no override. Lets new QBO companies in the master file work
    without a code change — they just don't get a custom label."""
    return DISPLAY_NAMES.get(qbo_company, qbo_company)


def _sort_key(label: str) -> tuple[int, str]:
    """Sort key that respects DISPLAY_ORDER, with unknown labels appended
    alphabetically just before YourTickets (which always goes last).

    Returns a (priority, label) tuple:
      - YourTickets → (2, '')               always last
      - Listed labels → (0, position)       in DISPLAY_ORDER index order
      - Unknown labels → (1, label)         alphabetically before YourTickets
    """
    if label == "YourTickets":
        return (2, "")
    try:
        return (0, f"{DISPLAY_ORDER.index(label):03d}")
    except ValueError:
        return (1, label)

# Columns dropped early (seat-level detail). Listed once so the schema is obvious.
SEAT_LEVEL_COLUMNS = [
    "Opponent/Performer", "Event Date", "Seat Section", "Seat Row", "Seats",
    "Ticket Cost Start", "Ticket Cost End", "Qty Start", "Qty End",
]

# Final output column order. "Original Company" is a helper used by the
# Bills/Expenses builders. Other helpers (e.g. "_display_label") are
# stripped before writing visible sheets.
FINAL_COLUMNS = [
    "Company", "Adjustment Date", "Vendor", "Team/Performer",
    "Total Adjustment", "Original Company",
]

# Aggregation key for the final collapse.
# Same (Company, Adjustment Date, Vendor, Team/Performer) rows get summed
# regardless of PO #, who entered them, or which seat lines they came from.
AGGREGATION_KEYS = [
    "Company", "Adjustment Date", "Vendor", "Team/Performer",
]

# Group keys for the FIRST collapse — one row per PO event — used by the
# cancellation override logic, which needs Cancelled and Total End.
GROUP_KEYS = [
    "Company", "Original Company", "PO #", "Adjustment Date", "Vendor",
    "Team/Performer", "Cancelled", "User",
]


def transform(
    df: pd.DataFrame,
    mapping: dict[str, str] | None = None,
) -> tuple[pd.DataFrame, dict]:
    """Run the full PO Cost Changes pipeline.

    Mirrors Power Query Section1.m → Table1. Step numbers below correspond
    to the M code's named steps.

    Args:
        df: raw input DataFrame matching the template's Table1 schema.
        mapping: optional override for the TicketVault→QBO company mapping.
                 If None, loads from Master_Mapping_List.xlsx via app.mapping.

    Returns:
        (cleaned_df, dropped_info) where dropped_info is:
            {"unmapped_companies": {"<name>": <row_count>, ...},
             "total_dropped_rows": int}
        The dropped info lets the UI flag accidental uploads of non-QBO companies.
    """
    if mapping is None:
        mapping = get_mapping()

    # 1. Source — caller already loaded the data.
    out = df.copy()

    # 2. Changed Type — coerce the columns that matter for math/comparison.
    out["PO #"] = pd.to_numeric(out["PO #"], errors="coerce").astype("Int64")
    out["Adjustment Date"] = pd.to_datetime(out["Adjustment Date"], errors="coerce")
    # Strip the time component — the source data has timestamps but every
    # row we care about is a per-day event, and the aggregation key
    # depends on dates matching exactly.
    out["Adjustment Date"] = out["Adjustment Date"].dt.normalize()
    for col in ["Ticket Cost Total Start", "Ticket Cost Total End",
                "Per Ticket Adjustment", "Total Adjustment"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
    for col in ["Company", "Vendor", "Team/Performer", "Cancelled", "User"]:
        if col in out.columns:
            out[col] = out[col].astype("string")
    # Cancelled feeds into boolean comparisons later; treat NaN/NA as empty string.
    if "Cancelled" in out.columns:
        out["Cancelled"] = out["Cancelled"].fillna("")

    # 3. Removed Columns — drop seat-level detail.
    out = out.drop(columns=[c for c in SEAT_LEVEL_COLUMNS if c in out.columns])

    # 4. Renamed Columns
    out = out.rename(columns={
        "Ticket Cost Total Start": "Total Start",
        "Ticket Cost Total End": "Total End",
    })

    # 4b. Vendor rename rule — vendors in TEAM_RENAME_VENDORS
    #     (Ticketmaster AM, Tickets.com, Ballpark) get replaced by the
    #     Team/Performer when that performer is one of the major-league
    #     teams. Other vendors are untouched. Applied before company
    #     mapping so the new vendor name participates in aggregation.
    if "Vendor" in out.columns and "Team/Performer" in out.columns:
        teams = get_teams()
        out["Vendor"] = out.apply(
            lambda r: rename_team_vendor(r["Vendor"], r["Team/Performer"], teams),
            axis=1,
        )

    # 5. Changed Type1 (Currency) — already numeric; skip.
    # 5b. Removed Columns1 — drop per-ticket adjustment.
    out = out.drop(columns=[c for c in ["Per Ticket Adjustment"] if c in out.columns])

    # 5c. Snapshot the original Company string (trimmed, casing preserved)
    #     before any mapping. The Bills/Expenses tabs need this for QBO entry.
    out["Original Company"] = out["Company"].map(
        lambda v: str(v).strip() if pd.notna(v) else v
    )

    # 6. Replaced Value — map TicketVault company names to QBO names.
    #    Case-insensitive: master keys are lowercased at load time; we
    #    lowercase the input before lookup. Output keeps the canonical
    #    QBO casing from the master file.
    out["Company"] = out["Company"].map(
        lambda v: mapping.get(str(v).strip().lower(), v) if pd.notna(v) else v
    )

    # 6b. Filter to QBO-mapped companies only.
    #     The master file's QBO Company column is the gating list — anything
    #     else is either an upload mistake or out of scope. We record what
    #     got dropped so the response surfaces it.
    #     Membership check is case-insensitive against the canonical QBO names,
    #     so an upload that uses the canonical name in any casing is preserved.
    canonical_qbo_names = set(mapping.values())
    canonical_qbo_lower = {n.lower() for n in canonical_qbo_names}
    is_mapped = out["Company"].map(
        lambda v: pd.notna(v) and str(v).strip().lower() in canonical_qbo_lower
    )
    dropped = out[~is_mapped]
    unmapped_counts: dict[str, int] = {}
    if not dropped.empty:
        for name, grp in dropped.groupby("Company", dropna=False):
            label = "(blank)" if pd.isna(name) else str(name)
            unmapped_counts[label] = int(len(grp))
    out = out[is_mapped].reset_index(drop=True)

    # 6c. Normalize Company to the canonical casing from the master, so that
    #     two inputs that differ only in casing (e.g. "Y&S Tickets" and
    #     "y&s tickets") collapse into a single output bucket.
    canonical_by_lower = {n.lower(): n for n in canonical_qbo_names}
    out["Company"] = out["Company"].map(
        lambda v: canonical_by_lower.get(str(v).strip().lower(), v) if pd.notna(v) else v
    )

    # 6d. Swap canonical QBO name for its display label (e.g. "YSKG Tickets" -> "YSKG").
    #     The Company column in all output files now uses the short label.
    out["Company"] = out["Company"].map(display_name)

    # 7. Grouped Rows — collapse seat lines into one PO event row.
    out = (
        out.groupby(GROUP_KEYS, dropna=False, as_index=False)
        .agg({
            "Total Start": "sum",
            "Total End": "sum",
            "Total Adjustment": "sum",
        })
    )

    # 8. Added Conditional Column — for cancellations, override adjustment
    #    to the full negative of Total End (reverses the booking).
    out["Total Adjustment"] = out.apply(
        lambda r: -r["Total End"] if r["Cancelled"] == "Yes" else r["Total Adjustment"],
        axis=1,
    )

    # 9. Filtered Rows — drop zero-impact rows (per the M code's behavior;
    #    further zero-sum aggregates are filtered again after step 12).
    out = out[out["Total Adjustment"] != 0].reset_index(drop=True)

    # 11. Cancelled is no longer needed beyond this point (step 8 has already
    #     applied its side-effect). It would otherwise carry into the
    #     aggregation key and split groups artificially.
    out = out.drop(columns=["Cancelled"])

    # 12. Final aggregation — collapse all rows with matching
    #     (Company, Adjustment Date, Vendor, Team/Performer) into one row,
    #     summing Total Adjustment. Drops PO #, Total Start, Total End, User.
    #     For Original Company, keep the most common value within each group
    #     (in case the same display-label bucket has rows that came in under
    #     slightly different original spellings — rare but possible).
    if not out.empty:
        out = (
            out.groupby(AGGREGATION_KEYS, dropna=False, as_index=False)
            .agg({
                "Total Adjustment": "sum",
                "Original Company": lambda s: s.mode().iat[0] if not s.mode().empty else s.iloc[0],
            })
        )

        # Filter zero-sum aggregates: separate +/− entries on the same key
        # can cancel out. They have no QBO impact, so drop them.
        out = out[out["Total Adjustment"] != 0].reset_index(drop=True)

    # 13. Sort for stable output: by display order (Company), then date.
    if not out.empty:
        out["_display_order"] = out["Company"].map(_sort_key)
        out = (
            out.sort_values(["_display_order", "Adjustment Date", "Vendor", "Team/Performer"],
                            kind="mergesort")
            .drop(columns=["_display_order"])
            .reset_index(drop=True)
        )

    cleaned = out[FINAL_COLUMNS]
    dropped_info = {
        "unmapped_companies": unmapped_counts,
        "total_dropped_rows": sum(unmapped_counts.values()),
    }
    return cleaned, dropped_info


def summarize_by_company(transformed: pd.DataFrame) -> dict:
    """Per-company row count and total cost change, plus grand totals.

    Returns:
        {
          "companies": [{"company": str, "rows": int, "total_adjustment": float}, ...],
          "totals": {"rows": int, "total_adjustment": float}
        }
    """
    if transformed.empty:
        return {"companies": [], "totals": {"rows": 0, "total_adjustment": 0.0}}

    grouped = (
        transformed.groupby("Company", dropna=False, as_index=False)
        .agg(rows=("Total Adjustment", "size"), total_adjustment=("Total Adjustment", "sum"))
    )
    # Sort by the same display order used everywhere else (YourTickets last).
    # NaN/blank Company values get pushed to the very end.
    grouped["_sort"] = grouped["Company"].map(
        lambda c: _sort_key(str(c)) if pd.notna(c) else (3, "")
    )
    grouped = grouped.sort_values("_sort").drop(columns=["_sort"])
    companies = [
        {
            "company": ("(blank)" if pd.isna(r["Company"]) else str(r["Company"])),
            "rows": int(r["rows"]),
            "total_adjustment": float(r["total_adjustment"]),
        }
        for _, r in grouped.iterrows()
    ]
    return {
        "companies": companies,
        "totals": {
            "rows": int(len(transformed)),
            "total_adjustment": float(transformed["Total Adjustment"].sum()),
        },
    }


# ---------------------------------------------------------------------------
# High-level orchestration for the web layer
# ---------------------------------------------------------------------------

import io
from datetime import date
from pathlib import Path
import logging

log = logging.getLogger(__name__)


def _ordinal(n: int) -> str:
    """1 -> '1st', 11 -> '11th', 22 -> '22nd'."""
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    return f"{n}{ {1:'st', 2:'nd', 3:'rd'}.get(n % 10, 'th') }"


def _format_date_range(dates: list[pd.Timestamp]) -> str:
    """e.g. 'May 1st thru May 3rd 2026' — same shape as the reference project."""
    valid = [d for d in dates if pd.notna(d)]
    if not valid:
        return date.today().strftime("%B %Y")
    valid = sorted(valid)

    def fmt(d):
        return f"{d.strftime('%B')} {_ordinal(d.day)}"

    if len(valid) == 1 or valid[0] == valid[-1]:
        return f"{fmt(valid[0])} {valid[0].strftime('%Y')}"
    return f"{fmt(valid[0])} thru {fmt(valid[-1])} {valid[-1].strftime('%Y')}"


def _read_one(content: bytes, filename: str) -> pd.DataFrame:
    """Read one uploaded file's bytes into a DataFrame.

    Recognizes .xlsx, .xlsm, .csv. For workbooks, tries the template's named
    sheet first; falls back to the first sheet.

    The real PO Cost Changes export has a two-row header: row 1 is category
    labels ("Ticket Cost", "Qty", ...) above merged cells, row 2 is the real
    field names. We detect this by checking whether "Company" appears in
    row 1's columns; if not, we re-read with header=1.
    """
    suffix = Path(filename).suffix.lower()
    buf = io.BytesIO(content)

    if suffix in (".xlsx", ".xlsm", ".xls"):
        try:
            df = pd.read_excel(buf, sheet_name="PO Cost Changes")
        except ValueError:
            buf.seek(0)
            df = pd.read_excel(buf)
        if "Company" not in df.columns:
            buf.seek(0)
            try:
                df = pd.read_excel(buf, sheet_name="PO Cost Changes", header=1)
            except ValueError:
                buf.seek(0)
                df = pd.read_excel(buf, header=1)
        return df

    if suffix == ".csv":
        df = pd.read_csv(buf)
        if "Company" not in df.columns:
            buf.seek(0)
            df = pd.read_csv(buf, header=1)
        return df

    raise ValueError(f"Unsupported file type for {filename!r}: {suffix}")


def _read_purchase_details(content: bytes, filename: str) -> pd.DataFrame:
    """Read one Purchase Details file. Same formats as PO Cost Changes,
    but the header is on row 1 (no merged category row above it). We still
    fall back to header=1 if 'PO #' isn't found in row 1, just in case."""
    suffix = Path(filename).suffix.lower()
    buf = io.BytesIO(content)
    if suffix in (".xlsx", ".xlsm", ".xls"):
        df = pd.read_excel(buf)
        if "PO #" not in df.columns:
            buf.seek(0)
            df = pd.read_excel(buf, header=1)
        return df
    if suffix == ".csv":
        df = pd.read_csv(buf)
        if "PO #" not in df.columns:
            buf.seek(0)
            df = pd.read_csv(buf, header=1)
        return df
    raise ValueError(f"Unsupported Purchase Details file type for {filename!r}: {suffix}")


def _collect_excluded_pos(
    pd_file_list: list[tuple[bytes, str]],
) -> tuple[set[int], int]:
    """Read all Purchase Details files, return the union of their PO numbers
    plus a total raw row count (for the response payload)."""
    if not pd_file_list:
        return set(), 0
    pos: set[int] = set()
    total = 0
    for content, filename in pd_file_list:
        df = _read_purchase_details(content, filename)
        if "PO #" not in df.columns:
            log.warning("Purchase Details file %s has no 'PO #' column; skipping.", filename)
            continue
        nums = pd.to_numeric(df["PO #"], errors="coerce").dropna().astype(int)
        pos.update(nums.tolist())
        total += len(df)
        log.info("Read Purchase Details %s: %d rows, %d distinct PO #", filename, len(df), nums.nunique())
    return pos, total


def _write_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    """Write df to a new sheet with the same styling as the reference project:
    blue header, alternating row fills, borders, frozen header, auto-filter,
    currency formatting on money columns, mm/dd/yyyy on Adjustment Date.

    Sheet names are sanitized to Excel's rules (≤31 chars, no `:\\/?*[]`).
    """
    safe = sheet_name[:31]
    for ch in r":\/?*[]":
        safe = safe.replace(ch, "_")

    ws = wb.create_sheet(safe)
    cols = list(df.columns)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    fill_odd  = PatternFill("solid", start_color="FFFFFF")
    fill_even = PatternFill("solid", start_color="EEF2FF")
    money_cols = {"Total Start", "Total End", "Total Adjustment", "Total"}

    for ri, row in enumerate(df.itertuples(index=False), 2):
        row_fill = fill_even if ri % 2 == 0 else fill_odd
        for ci, val in enumerate(row, 1):
            col_name = cols[ci - 1]
            # Cast pandas NA / NaT to None so openpyxl writes empty cells.
            if pd.isna(val):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            cell.fill = row_fill
            if col_name in ("Adjustment Date", "Date") and val is not None:
                cell.number_format = "mm/dd/yyyy"
            elif col_name in money_cols and val is not None:
                cell.number_format = '"$"#,##0.00;[Red]"-$"#,##0.00'

    # Column widths sized to content (capped at 55).
    for ci, col in enumerate(cols, 1):
        max_len = len(str(col))
        for row in df.itertuples(index=False):
            v = row[ci - 1]
            max_len = max(max_len, 0 if v is None or (isinstance(v, float) and pd.isna(v)) else len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 55)

    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions


def _build_bills_and_expenses(
    cleaned: pd.DataFrame,
    open_vendors: set[str] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build the Bills and Expenses ledger views.

    Bills: one row per PO whose Total Adjustment > 0. Total stays positive.

    Expenses: two rows per PO whose Total Adjustment < 0.
      - Line A: Category = 'Inventory Asset', Total = Total Adjustment (negative)
      - Line B: Category = '<Vendor> (TC)' OR 'Due from Vendors - Open' if vendor
                is on the open-vendors list. Total = -Total Adjustment (positive),
                so the pair sums to zero.

    Both share the same column shape and a single global 'Expense #' counter
    incrementing across the whole result (Bills + Expenses interleaved by
    display order, then by PO #).
    """
    if open_vendors is None:
        open_vendors = get_open_vendors()

    ledger_cols = [
        "Company", "Account", "Date", "Category", "Expense #",
        "Vendor", "Memo", "Description", "Total",
    ]

    # Empty input → empty frames with the right shape, so downstream writers
    # can still render headers.
    if cleaned.empty:
        empty = pd.DataFrame(columns=ledger_cols)
        return empty.copy(), empty.copy()

    # Order events by display order, then date and vendor, so the global
    # Expense # is deterministic and groups each company's rows together.
    df = cleaned.copy()
    df["_display_order"] = df["Company"].map(_sort_key)
    df = df.sort_values(
        ["_display_order", "Adjustment Date", "Vendor", "Team/Performer"],
        kind="mergesort",
    ).reset_index(drop=True)
    df["Expense #"] = range(1, len(df) + 1)

    def _memo(row) -> str:
        team = row["Team/Performer"] if pd.notna(row["Team/Performer"]) else ""
        orig = row["Original Company"] if pd.notna(row["Original Company"]) else ""
        return f"{team} - Cost Changes ({orig})"

    df["_memo"] = df.apply(_memo, axis=1)

    # Bills: positive adjustments → single row each.
    bills_src = df[df["Total Adjustment"] > 0].copy()
    bills = pd.DataFrame({
        "Company":     bills_src["Original Company"],
        "Account":     "Clearing Account",
        "Date":        bills_src["Adjustment Date"],
        "Category":    "Inventory Asset",
        "Expense #":   bills_src["Expense #"],
        "Vendor":      bills_src["Vendor"],
        "Memo":        bills_src["_memo"],
        "Description": bills_src["_memo"],
        "Total":       bills_src["Total Adjustment"],
    })

    # Expenses: negative adjustments → two rows each, summing to zero.
    exp_src = df[df["Total Adjustment"] < 0].copy()
    if not exp_src.empty:
        # Line A — Inventory Asset, negative
        line_a = pd.DataFrame({
            "Company":     exp_src["Original Company"],
            "Account":     "Clearing Account",
            "Date":        exp_src["Adjustment Date"],
            "Category":    "Inventory Asset",
            "Expense #":   exp_src["Expense #"],
            "Vendor":      exp_src["Vendor"],
            "Memo":        exp_src["_memo"],
            "Description": exp_src["_memo"],
            "Total":       exp_src["Total Adjustment"],   # already negative
        })
        # Line B — Vendor (TC) or Due from Vendors - Open, positive offset
        line_b = pd.DataFrame({
            "Company":     exp_src["Original Company"],
            "Account":     "Clearing Account",
            "Date":        exp_src["Adjustment Date"],
            "Category":    exp_src["Vendor"].map(lambda v: offset_category(v, open_vendors)),
            "Expense #":   exp_src["Expense #"],
            "Vendor":      exp_src["Vendor"],
            "Memo":        exp_src["_memo"],
            "Description": exp_src["_memo"],
            "Total":       -exp_src["Total Adjustment"],  # positive (flips sign)
        })
        # Interleave A,B,A,B,... by sorting on (Expense #, line_order)
        line_a["_line"] = 0
        line_b["_line"] = 1
        expenses = (
            pd.concat([line_a, line_b], ignore_index=True)
            .sort_values(["Expense #", "_line"], kind="mergesort")
            .drop(columns=["_line"])
            .reset_index(drop=True)
        )
    else:
        expenses = pd.DataFrame(columns=ledger_cols)

    # Attach display label as a hidden helper column on Expenses so the
    # per-company tab filtering can use it. Map via Expense #.
    display_by_expense = dict(zip(df["Expense #"], df["Company"]))
    if not expenses.empty:
        expenses["_display_label"] = expenses["Expense #"].map(display_by_expense)
    if not bills.empty:
        bills["_display_label"] = bills["Expense #"].map(display_by_expense)

    return bills[ledger_cols + (["_display_label"] if "_display_label" in bills else [])], \
           expenses[ledger_cols + (["_display_label"] if "_display_label" in expenses else [])]


def _apply_cancelled_override_raw(df: pd.DataFrame) -> pd.DataFrame:
    """Apply the Cancelled override to a RAW DataFrame (pre-pipeline shape).

    When Cancelled == "Yes", set Total Adjustment = -(Ticket Cost Total End).
    This mirrors step 8 of transform(), but for raw rows so the Source Data
    and Excluded tabs show numbers that reconcile cleanly to the Combined
    ledger (which uses the same override internally).

    Returns a new DataFrame with the column updated; the original is untouched.
    """
    if "Cancelled" not in df.columns or "Total Adjustment" not in df.columns:
        return df
    out = df.copy()
    is_cancelled = out["Cancelled"].astype("string").str.strip().str.lower().eq("yes")
    if "Ticket Cost Total End" in out.columns:
        end = pd.to_numeric(out["Ticket Cost Total End"], errors="coerce")
        out.loc[is_cancelled, "Total Adjustment"] = -end[is_cancelled]
    return out


def _write_summary_sheet(wb, combined_ledger: pd.DataFrame) -> None:
    """Write the 'Summary' sheet: a pivot-style outline view of the Combined
    ledger, grouped Company > Vendor > Description, with one row per (Company,
    Vendor, Description, Date). Subtotal rows after each Company group; grand
    total at the bottom. Outline groups are expanded by default; the user can
    collapse with Excel's outline +/− buttons in the left margin.

    Matches the layout in the user's screenshot: blue header band, currency
    formatting on Total (red for negatives), date in mm/dd/yyyy.
    """
    ws = wb.create_sheet("Summary")

    # ── Shared styling ──────────────────────────────────────────────────────
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    company_font = Font(name="Arial", bold=True, size=10)
    company_fill = PatternFill("solid", start_color="D9E2F3")
    subtotal_font = Font(name="Arial", bold=True, color="C00000", size=10)
    subtotal_fill = PatternFill("solid", start_color="EEF2FF")
    grand_font = Font(name="Arial", bold=True, size=11)
    body_font = Font(name="Arial", size=10)
    money_fmt = '"$"#,##0.00;[Red]"-$"#,##0.00'

    # ── Empty data → just write headers and bail ────────────────────────────
    if combined_ledger.empty:
        for ci, h in enumerate(["Company", "Vendor", "Description", "Date", "Total"], 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        ws.freeze_panes = "A2"
        return

    # ── Aggregate to (Company, Vendor, Description, Date) ──────────────────
    df = combined_ledger.copy()
    df["Total"] = pd.to_numeric(df["Total"], errors="coerce").fillna(0.0)
    grouped = (
        df.groupby(["Company", "Vendor", "Description", "Date"], dropna=False, as_index=False)
        ["Total"].sum()
    )
    # Sort: company by the same DISPLAY_ORDER used everywhere else; everything
    # else alphabetically/chronologically.
    grouped["_co_order"] = grouped["Company"].map(_sort_key)
    grouped = grouped.sort_values(
        ["_co_order", "Vendor", "Description", "Date"], kind="mergesort"
    ).drop(columns=["_co_order"]).reset_index(drop=True)

    # ── Title + header row ──────────────────────────────────────────────────
    # Row 1 — "Sum of Total" label (matches the screenshot's banner)
    ws.cell(row=1, column=1, value="Sum of Total").font = Font(name="Arial", bold=True, size=11)
    # Row 2 — column headers
    headers = ["Company", "Vendor", "Description", "Date", "Total"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border

    # ── Body: one row per (Company, Vendor, Description, Date) with
    #    Company shown only on the first row of its group + a subtotal row. ──
    row = 3
    company_row_ranges: list[tuple[str, int, int]] = []  # (company, start, end_data)
    for company, co_grp in grouped.groupby("Company", sort=False, dropna=False):
        co_start = row
        first_row_of_company = True
        for _, r in co_grp.iterrows():
            # Company label — only on the first row of the group
            ws.cell(row=row, column=1, value=str(company) if first_row_of_company else None)
            ws.cell(row=row, column=2, value=str(r["Vendor"]) if pd.notna(r["Vendor"]) else None)
            ws.cell(row=row, column=3, value=str(r["Description"]) if pd.notna(r["Description"]) else None)
            date_val = r["Date"] if pd.notna(r["Date"]) else None
            ws.cell(row=row, column=4, value=date_val)
            ws.cell(row=row, column=5, value=float(r["Total"]))

            # Styling
            ws.cell(row=row, column=4).number_format = "mm/dd/yyyy"
            ws.cell(row=row, column=5).number_format = money_fmt
            if first_row_of_company:
                ws.cell(row=row, column=1).font = company_font
                ws.cell(row=row, column=1).fill = company_fill
            for ci in range(1, 6):
                cell = ws.cell(row=row, column=ci)
                if cell.font.name is None:
                    cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(vertical="center")

            first_row_of_company = False
            row += 1

        # Subtotal row for this Company
        ws.cell(row=row, column=1, value=f"{company} Total")
        ws.cell(row=row, column=5, value=f"=SUBTOTAL(9,E{co_start}:E{row - 1})")
        for ci in range(1, 6):
            cell = ws.cell(row=row, column=ci)
            cell.font = subtotal_font; cell.fill = subtotal_fill
            cell.border = border; cell.alignment = Alignment(vertical="center")
        ws.cell(row=row, column=5).number_format = money_fmt

        company_row_ranges.append((str(company), co_start, row - 1))
        row += 1

    # ── Grand total at the bottom ──────────────────────────────────────────
    grand_row = row
    ws.cell(row=grand_row, column=1, value="Grand Total")
    if company_row_ranges:
        first_start = company_row_ranges[0][1]
        last_end = company_row_ranges[-1][2]
        # Use SUM (not SUBTOTAL) since we want it to ignore the subtotal rows.
        # Easiest way: sum the per-company subtotal cells.
        subtotal_cells = [f"E{end + 1}" for _, _, end in company_row_ranges]
        ws.cell(row=grand_row, column=5, value=f"={'+'.join(subtotal_cells)}")
    else:
        ws.cell(row=grand_row, column=5, value=0)
    for ci in range(1, 6):
        cell = ws.cell(row=grand_row, column=ci)
        cell.font = grand_font; cell.fill = company_fill
        cell.border = Border(left=thin, right=thin,
                             top=Side(style="medium", color="000000"), bottom=thin)
        cell.alignment = Alignment(vertical="center")
    ws.cell(row=grand_row, column=5).number_format = money_fmt

    # ── Outline groups for collapse/expand on Company column ───────────────
    # Excel outline: data rows of each company are grouped one level deep;
    # collapsing hides everything except the subtotal row. Default expanded.
    for _, start, end in company_row_ranges:
        if end >= start:
            for r in range(start, end + 1):
                ws.row_dimensions[r].outline_level = 1
                ws.row_dimensions[r].hidden = False
    ws.sheet_properties.outlinePr.summaryBelow = True

    # ── Column widths and freeze pane ──────────────────────────────────────
    widths = {1: 24, 2: 24, 3: 60, 4: 12, 5: 14}
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A3"


def _build_combined_workbook(
    source_df: pd.DataFrame,
    bills_df: pd.DataFrame,
    expenses_df: pd.DataFrame,
    all_company_labels: list[str],
    excluded_df: pd.DataFrame | None = None,
) -> bytes:
    """Build the multi-sheet combined workbook:
       - 'Source Data' — the raw merged upload, untouched
       - 'Combined'    — one row per aggregated event (ledger format)
       - 'Bills'       — positive-adjustment events (one row each)
       - One tab per company — the company's Expense pairs (debit/credit)
       - 'Excluded'    — raw rows filtered out via Purchase Details match
                         (last tab; only shown if any rows were excluded)
       - Empty per-company tabs get a red tab color
    """
    drop_helper = lambda d: d.drop(columns=[c for c in ["_display_label"] if c in d.columns])
    bills_visible = drop_helper(bills_df)
    expenses_visible = drop_helper(expenses_df)

    # 'Combined' = every aggregated event as a single ledger row:
    #   bills as-is + the negative (Inventory Asset) leg of expenses,
    #   ordered by Expense #.
    if not expenses_visible.empty:
        expense_singles = expenses_visible[
            expenses_visible["Category"] == "Inventory Asset"
        ]
    else:
        expense_singles = expenses_visible
    combined_ledger = pd.concat(
        [bills_visible, expense_singles], ignore_index=True
    ).sort_values("Expense #", kind="mergesort").reset_index(drop=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary first — pivot-style outline of the Combined ledger.
    _write_summary_sheet(wb, combined_ledger)

    # Source data
    _write_sheet(wb, "Source Data", source_df)

    # Excluded tab — right after Source Data so it's adjacent to the raw view.
    # Only added when Purchase Details were uploaded and at least one row got
    # filtered out. Shape matches Source Data so the user sees exactly which
    # raw rows were removed.
    if excluded_df is not None and len(excluded_df) > 0:
        _write_sheet(wb, "Excluded", excluded_df.reset_index(drop=True))

    # Combined ledger
    _write_sheet(wb, "Combined", combined_ledger)

    # Bills tab (all positive events, all companies)
    _write_sheet(wb, "Bills", bills_visible)

    # One tab per company with Expense pairs
    for label in all_company_labels:
        if "_display_label" in expenses_df.columns:
            tab_df = (
                expenses_df[expenses_df["_display_label"] == label]
                .drop(columns=["_display_label"])
                .reset_index(drop=True)
            )
        else:
            tab_df = expenses_df.iloc[0:0]
        _write_sheet(wb, label, tab_df)
        if len(tab_df) == 0:
            wb[label[:31]].sheet_properties.tabColor = "FF0000"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_company_workbook(
    label: str,
    bills_df: pd.DataFrame,
    expenses_df: pd.DataFrame,
) -> bytes:
    """Per-company download file: two sheets (Bills + Expenses)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if "_display_label" in bills_df.columns:
        b = bills_df[bills_df["_display_label"] == label].drop(columns=["_display_label"])
    else:
        b = bills_df.iloc[0:0]
    if "_display_label" in expenses_df.columns:
        e = expenses_df[expenses_df["_display_label"] == label].drop(columns=["_display_label"])
    else:
        e = expenses_df.iloc[0:0]

    _write_sheet(wb, "Bills", b.reset_index(drop=True))
    _write_sheet(wb, "Expenses", e.reset_index(drop=True))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_single_sheet_xlsx(df: pd.DataFrame, sheet_name: str = "PO Cost Changes") -> bytes:
    """Single-sheet styled workbook — kept for compatibility / future use."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, sheet_name, df)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def process_files(
    file_list: list[tuple[bytes, str]],
    pd_file_list: list[tuple[bytes, str]] | None = None,
) -> dict:
    """Process N uploaded files into the bundle the Flask app needs.

    Args:
        file_list:    PO Cost Changes uploads — list of (bytes, filename) tuples.
        pd_file_list: optional Purchase Details uploads. Any PO # appearing in
                      these files is excluded from the output (except Source Data).

    Returns:
        {
          "date_range": "May 1st thru May 3rd 2026",
          "combined": <xlsx bytes>,                        # multi-tab workbook
          "companies": {label: <xlsx bytes>, ...},          # only labels with data
          "all_companies": [...],                           # all QBO labels, sorted
          "stats": { "Combined": {...}, "<label>": {...}, ... },
          "dropped": {"unmapped_companies": {...}, "total_dropped_rows": N},
          "excluded": {"po_count": N, "row_count": N},      # how many got filtered
        }
    """
    if not file_list:
        raise ValueError("No files provided")

    # 1) Read + merge all PO Cost Changes uploads.
    frames = []
    for content, filename in file_list:
        df = _read_one(content, filename)
        frames.append(df)
        log.info("Read %s: %d rows", filename, len(df))
    merged = pd.concat(frames, ignore_index=True, sort=False)

    # 1a) Filter out any company that isn't in the QBO master mapping.
    #     Doing this on the raw merge means Source Data and Excluded tabs
    #     only show in-scope rows, and the math reconciles cleanly:
    #     Source Data total − Excluded total = Combined total.
    mapping_now = get_mapping()
    valid_lower = set(mapping_now.keys()) | {v.lower() for v in mapping_now.values()}
    raw_company = merged["Company"].astype("string").str.strip().str.lower()
    in_scope_mask = raw_company.isin(valid_lower)
    out_of_scope_counts: dict[str, int] = {}
    if (~in_scope_mask).any():
        out_of_scope = merged.loc[~in_scope_mask, "Company"].astype("string").fillna("(blank)")
        out_of_scope_counts = out_of_scope.value_counts().to_dict()
        log.info(
            "Ignored %d rows from %d out-of-scope compan%s: %s",
            int((~in_scope_mask).sum()),
            len(out_of_scope_counts),
            "y" if len(out_of_scope_counts) == 1 else "ies",
            ", ".join(f"{k} ({v})" for k, v in out_of_scope_counts.items()),
        )
    merged = merged[in_scope_mask].reset_index(drop=True)

    # 1b) Determine which rows to exclude from output. Two independent reasons:
    #     (a) PO # appears in the uploaded Purchase Details files
    #     (b) the row's "Remove" column is marked with X (case-insensitive)
    #     A row excluded for EITHER reason goes to the Excluded tab exactly once.
    #     Source Data still shows everything.
    excluded_pos, _pd_total_rows = _collect_excluded_pos(pd_file_list or [])

    po_col = pd.to_numeric(merged["PO #"], errors="coerce").astype("Int64")
    po_match_mask = po_col.isin(excluded_pos) if excluded_pos else pd.Series(False, index=merged.index)

    # Locate the "Remove" column case-insensitively (header may be Remove/remove/REMOVE).
    remove_col = next(
        (c for c in merged.columns if str(c).strip().lower() == "remove"),
        None,
    )
    if remove_col is not None:
        remove_mask = merged[remove_col].astype("string").str.strip().str.lower().eq("x")
        remove_mask = remove_mask.fillna(False)
    else:
        remove_mask = pd.Series(False, index=merged.index)

    excluded_mask = po_match_mask | remove_mask

    if excluded_mask.any():
        excluded_raw = merged[excluded_mask].reset_index(drop=True)
        merged_for_pipeline = merged[~excluded_mask].reset_index(drop=True)
        log.info(
            "Excluded %d rows (%d via Purchase Details PO match, %d via Remove=X, "
            "%d caught by both) covering %d PO #s",
            int(excluded_mask.sum()),
            int(po_match_mask.sum()),
            int(remove_mask.sum()),
            int((po_match_mask & remove_mask).sum()),
            excluded_raw["PO #"].nunique() if len(excluded_raw) else 0,
        )
    else:
        excluded_raw = merged.iloc[0:0].copy()
        merged_for_pipeline = merged

    # 1c) Apply the Cancelled override to the Source Data and Excluded raw
    #     views, so their Total Adjustment column reconciles with the
    #     Combined ledger (which already applies the same override inside
    #     transform()). Doesn't touch the pipeline path — transform() applies
    #     its own override there.
    source_data_view = _apply_cancelled_override_raw(merged)
    excluded_view = _apply_cancelled_override_raw(excluded_raw)

    # 2) Run the canonical pipeline on the kept rows.
    cleaned, dropped = transform(merged_for_pipeline)

    # 3) Date range from the Adjustment Date column.
    date_range_str = _format_date_range(
        [d for d in cleaned["Adjustment Date"].tolist() if pd.notna(d)]
    )

    # 4) Per-company DataFrames (only ones with data), ordered by display
    #    order so the dict iteration matches the UI grid and tab order.
    grouped_dfs: dict[str, pd.DataFrame] = {}
    for company, grp in cleaned.groupby("Company", dropna=False):
        if pd.isna(company):
            continue
        grouped_dfs[str(company)] = grp.reset_index(drop=True)
    company_dfs: dict[str, pd.DataFrame] = {
        k: grouped_dfs[k] for k in sorted(grouped_dfs.keys(), key=_sort_key)
    }

    # 5) All canonical QBO companies from the master file, translated to
    #    display labels, ordered to match the Purchase Details processor's
    #    tab order (YourTickets last). Used by the UI to render the full grid
    #    and by the combined workbook for tab order.
    all_companies = sorted(
        {display_name(n) for n in get_mapping().values()},
        key=_sort_key,
    )

    # 6) Stats block matching the reference's shape.
    #    Note: reference uses "total_cost"; we keep that key name so the UI
    #    template doesn't need editing. The value is sum of Total Adjustment.
    stats: dict[str, dict] = {
        "Combined": {
            "rows": int(len(cleaned)),
            "total_cost": round(float(cleaned["Total Adjustment"].sum()), 2) if len(cleaned) else 0.0,
        }
    }
    for name in all_companies:
        cdf = company_dfs.get(name)
        if cdf is not None and len(cdf) > 0:
            stats[name] = {
                "rows": int(len(cdf)),
                "total_cost": round(float(cdf["Total Adjustment"].sum()), 2),
            }
        else:
            stats[name] = {"rows": 0, "total_cost": 0.0}

    # 7) Build Bills (one row per positive event) and Expenses (debit/credit
    #    pairs per negative event), with a single global Expense # counter.
    bills_df, expenses_df = _build_bills_and_expenses(cleaned)

    # 8) Serialize everything to xlsx bytes.
    combined_bytes = _build_combined_workbook(
        source_data_view, bills_df, expenses_df, all_companies, excluded_view,
    )
    company_files = {
        name: _build_company_workbook(name, bills_df, expenses_df)
        for name in company_dfs.keys()
    }

    # Excluded $ total — same Cancelled-override convention as the Combined
    # ledger, so Source total − Excluded total = Combined total.
    excluded_total = (
        float(pd.to_numeric(excluded_view["Total Adjustment"], errors="coerce").fillna(0).sum())
        if len(excluded_view) else 0.0
    )

    return {
        "date_range": date_range_str,
        "combined": combined_bytes,
        "companies": company_files,
        "all_companies": all_companies,
        "stats": stats,
        "dropped": dropped,
        "excluded": {
            "po_count": int(excluded_raw["PO #"].nunique()) if len(excluded_raw) else 0,
            "row_count": int(len(excluded_raw)),
            "total_adjustment": round(excluded_total, 2),
        },
        "ignored_companies": out_of_scope_counts,
    }
