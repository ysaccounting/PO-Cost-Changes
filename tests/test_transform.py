"""Tests for the transform pipeline and the company summary."""
from __future__ import annotations

import pandas as pd
import pytest

from processor import display_name, summarize_by_company, transform


# Mapping fixture matches what's in the real master file for these names.
# Keys are lowercased — load_mapping() lowercases them at load time for
# case-insensitive lookup. Values keep canonical casing.
TEST_MAPPING = {
    "ysa":             "YS Asher Tickets",
    "ysa 2":           "YS Asher Tickets",
    "ysa 3":           "YS Asher Tickets",
    "ys tickets":      "Y&S Tickets",
    "ys-seatgeek":     "Y&S Tickets",
    "ys-seatgeek2":    "Y&S Tickets",
    "jacks ys":        "YS Chase Tickets",
    "yoni levine":     "YS Levine Tickets",
    "gk llc":          "YSKG Tickets",
}


def _raw_row(**overrides):
    base = {
        "Company": "YSA 2",
        "PO #": 1001,
        "Adjustment Date": pd.Timestamp("2026-01-15"),
        "Vendor": "Vendor A",
        "Team/Performer": "Bills",
        "Opponent/Performer": "Dolphins",
        "Event Date": pd.Timestamp("2026-02-01"),
        "Seat Section": "100", "Seat Row": "A", "Seats": "1-2",
        "Ticket Cost Start": 100.0, "Ticket Cost End": 120.0,
        "Qty Start": 2, "Qty End": 2,
        "Ticket Cost Total Start": 200.0,
        "Ticket Cost Total End": 240.0,
        "Per Ticket Adjustment": 20.0,
        "Total Adjustment": 40.0,
        "Cancelled": "",
        "User": "alice",
    }
    base.update(overrides)
    return base


def test_master_mapping_collapses_ysa_variants_to_ysa_label():
    df = pd.DataFrame([
        _raw_row(Company="YSA",   **{"PO #": 1}),
        _raw_row(Company="YSA 2", **{"PO #": 2}),
        _raw_row(Company="YSA 3", **{"PO #": 3}),
    ])
    out, dropped = transform(df, mapping=TEST_MAPPING)
    # "YS Asher Tickets" (QBO) → "YSA" (display label)
    assert set(out["Company"].unique()) == {"YSA"}
    assert dropped["total_dropped_rows"] == 0


def test_unmapped_company_is_dropped_and_reported():
    df = pd.DataFrame([
        _raw_row(Company="YSA 2",            **{"PO #": 1, "Total Adjustment": 10.0}),
        _raw_row(Company="Some Random Co",   **{"PO #": 2, "Total Adjustment": 99.0}),
        _raw_row(Company="Bearhawk - Aaron", **{"PO #": 3, "Total Adjustment": 50.0}),
        _raw_row(Company="Bearhawk - Aaron", **{"PO #": 4, "Total Adjustment": 50.0}),
    ])
    out, dropped = transform(df, mapping=TEST_MAPPING)
    assert list(out["Company"]) == ["YSA"]
    assert dropped["total_dropped_rows"] == 3
    assert dropped["unmapped_companies"] == {
        "Some Random Co": 1,
        "Bearhawk - Aaron": 2,
    }


def test_canonical_qbo_name_in_input_passes_through_as_display_label():
    """If the input already uses a QBO name, it should not be dropped,
    and the output should use the display label."""
    df = pd.DataFrame([
        _raw_row(Company="YS Asher Tickets", **{"PO #": 1, "Total Adjustment": 10.0}),
    ])
    out, dropped = transform(df, mapping=TEST_MAPPING)
    assert len(out) == 1
    assert out.iloc[0]["Company"] == "YSA"   # display label, not QBO name
    assert dropped["total_dropped_rows"] == 0


def test_lookup_is_case_insensitive_with_display_label_output():
    """Real data has 'YS-Seatgeek2' but master had 'YS-SeatGeek2'.
    Different casings should still map; output should use the display label.
    Uses distinct Team/Performer per row so aggregation doesn't collapse them."""
    df = pd.DataFrame([
        _raw_row(Company="YS-Seatgeek2",  **{"PO #": 1, "Total Adjustment": 10.0, "Team/Performer": "A"}),
        _raw_row(Company="ys-seatgeek2",  **{"PO #": 2, "Total Adjustment": 20.0, "Team/Performer": "B"}),
        _raw_row(Company="YS-SEATGEEK2",  **{"PO #": 3, "Total Adjustment": 30.0, "Team/Performer": "C"}),
        _raw_row(Company="y&s tickets",   **{"PO #": 4, "Total Adjustment": 40.0, "Team/Performer": "D"}),
    ])
    out, dropped = transform(df, mapping=TEST_MAPPING)
    assert len(out) == 4
    assert dropped["total_dropped_rows"] == 0
    # All should collapse to "Y&S" (display label for Y&S Tickets).
    assert set(out["Company"].unique()) == {"Y&S"}


def test_blank_company_is_dropped_and_reported_as_blank():
    df = pd.DataFrame([
        _raw_row(Company="YSA",  **{"PO #": 1, "Total Adjustment": 10.0}),
        _raw_row(Company=None,   **{"PO #": 2, "Total Adjustment": 20.0}),
    ])
    out, dropped = transform(df, mapping=TEST_MAPPING)
    assert len(out) == 1
    assert dropped["total_dropped_rows"] == 1
    assert "(blank)" in dropped["unmapped_companies"]


def test_seat_lines_collapse_to_single_po_row():
    """Two seat lines on the same PO and same group key collapse to one row."""
    df = pd.DataFrame([
        _raw_row(Company="YSA 2", **{"PO #": 9001, "Seats": "1",
                    "Ticket Cost Total Start": 100.0,
                    "Ticket Cost Total End": 110.0,
                    "Total Adjustment": 10.0}),
        _raw_row(Company="YSA 2", **{"PO #": 9001, "Seats": "2",
                    "Ticket Cost Total Start": 200.0,
                    "Ticket Cost Total End": 220.0,
                    "Total Adjustment": 20.0}),
    ])
    out, _ = transform(df, mapping=TEST_MAPPING)
    assert len(out) == 1
    assert out.iloc[0]["Total Adjustment"] == 30.0


def test_cancellation_overrides_adjustment_to_minus_total_end():
    """If Cancelled=Yes, the row's contribution becomes -Total End."""
    df = pd.DataFrame([
        _raw_row(Company="YSA", **{
            "PO #": 7000, "Cancelled": "Yes",
            "Ticket Cost Total Start": 500.0,
            "Ticket Cost Total End": 500.0,
            "Total Adjustment": 0.0,
        }),
    ])
    out, _ = transform(df, mapping=TEST_MAPPING)
    assert len(out) == 1
    assert out.iloc[0]["Total Adjustment"] == -500.0


def test_zero_adjustment_rows_filtered():
    """Rows with no net adjustment should be filtered out."""
    df = pd.DataFrame([
        _raw_row(Company="YSA", **{"PO #": 100, "Total Adjustment": 0.0,
                    "Ticket Cost Total End": 100.0}),
        _raw_row(Company="YSA", **{"PO #": 101, "Total Adjustment": 5.0,
                    "Vendor": "Other", "Team/Performer": "Other"}),
    ])
    out, _ = transform(df, mapping=TEST_MAPPING)
    assert len(out) == 1
    assert out.iloc[0]["Total Adjustment"] == 5.0


def test_aggregation_ignores_time_component_of_adjustment_date():
    """The real CSVs have timestamps on Adjustment Date. Same-day rows with
    different times must still aggregate together."""
    df = pd.DataFrame([
        _raw_row(Company="YSA", **{"PO #": 1, "Total Adjustment": 100.0,
                    "Adjustment Date": pd.Timestamp("2026-05-20 00:01:00"),
                    "Vendor": "Ticketmaster", "Team/Performer": "Atlanta Hawks"}),
        _raw_row(Company="YSA", **{"PO #": 2, "Total Adjustment": 250.0,
                    "Adjustment Date": pd.Timestamp("2026-05-20 14:30:00"),
                    "Vendor": "Ticketmaster", "Team/Performer": "Atlanta Hawks"}),
        _raw_row(Company="YSA", **{"PO #": 3, "Total Adjustment":  50.0,
                    "Adjustment Date": pd.Timestamp("2026-05-20 23:59:59"),
                    "Vendor": "Ticketmaster", "Team/Performer": "Atlanta Hawks"}),
    ])
    out, _ = transform(df, mapping=TEST_MAPPING)
    assert len(out) == 1
    assert out.iloc[0]["Total Adjustment"] == 400.0


def test_aggregation_collapses_same_company_date_vendor_team():
    """Multiple POs with the same (Company, Date, Vendor, Team/Performer)
    collapse into one row with summed Total Adjustment."""
    df = pd.DataFrame([
        _raw_row(Company="YSA", **{"PO #": 1, "Total Adjustment": 10.0,
                    "Vendor": "Ticketmaster", "Team/Performer": "Bills"}),
        _raw_row(Company="YSA", **{"PO #": 2, "Total Adjustment": 20.0,
                    "Vendor": "Ticketmaster", "Team/Performer": "Bills"}),
        _raw_row(Company="YSA", **{"PO #": 3, "Total Adjustment":  5.0,
                    "Vendor": "Ticketmaster", "Team/Performer": "Mets"}),
    ])
    out, _ = transform(df, mapping=TEST_MAPPING)
    # YSA + Ticketmaster + Bills = 30 (PO 1 + PO 2), YSA + Ticketmaster + Mets = 5
    assert len(out) == 2
    bills_row = out[out["Team/Performer"] == "Bills"].iloc[0]
    mets_row  = out[out["Team/Performer"] == "Mets"].iloc[0]
    assert bills_row["Total Adjustment"] == 30.0
    assert mets_row["Total Adjustment"] == 5.0


def test_aggregation_drops_zero_sum_groups():
    """If +/- cancel out within a group, the row should be dropped."""
    df = pd.DataFrame([
        _raw_row(Company="YSA", **{"PO #": 1, "Total Adjustment":  50.0,
                    "Vendor": "TM", "Team/Performer": "Bills"}),
        _raw_row(Company="YSA", **{"PO #": 2, "Total Adjustment": -50.0,
                    "Vendor": "TM", "Team/Performer": "Bills"}),
    ])
    out, _ = transform(df, mapping=TEST_MAPPING)
    assert len(out) == 0


def test_final_column_order():
    df = pd.DataFrame([_raw_row()])
    out, _ = transform(df, mapping=TEST_MAPPING)
    assert list(out.columns) == [
        "Company", "Adjustment Date", "Vendor", "Team/Performer",
        "Total Adjustment", "Original Company",
    ]


def test_summarize_by_company_groups_counts_and_totals():
    df = pd.DataFrame([
        _raw_row(Company="YSA",        **{"PO #": 1, "Total Adjustment":  10.0, "Team/Performer": "A"}),
        _raw_row(Company="YSA 2",      **{"PO #": 2, "Total Adjustment":  15.0, "Team/Performer": "B"}),
        _raw_row(Company="YSA 3",      **{"PO #": 3, "Total Adjustment": -25.0, "Team/Performer": "C"}),
        _raw_row(Company="YS Tickets", **{"PO #": 4, "Total Adjustment": 100.0, "Team/Performer": "D"}),
    ])
    out, _ = transform(df, mapping=TEST_MAPPING)
    s = summarize_by_company(out)

    by_co = {c["company"]: c for c in s["companies"]}
    # YSA Tickets (QBO) → "YSA" display label
    assert by_co["YSA"]["rows"] == 3
    assert by_co["YSA"]["total_adjustment"] == 0.0
    # Y&S Tickets (QBO) → "Y&S" display label
    assert by_co["Y&S"]["rows"] == 1
    assert by_co["Y&S"]["total_adjustment"] == 100.0

    assert s["totals"]["rows"] == 4
    assert s["totals"]["total_adjustment"] == 100.0


def test_summarize_empty_dataframe():
    df = pd.DataFrame(columns=[
        "Company","PO #","Adjustment Date","Vendor","Team/Performer",
        "Total Start","Total End","Total Adjustment","Cancelled","User"
    ])
    s = summarize_by_company(df)
    assert s["companies"] == []
    assert s["totals"]["rows"] == 0
    assert s["totals"]["total_adjustment"] == 0.0


def test_display_name_maps_known_qbo_names():
    assert display_name("YSKG Tickets") == "YSKG"
    assert display_name("YS Chase Tickets") == "Chase (Jacks)"
    assert display_name("The Ticket Guy LLC") == "The Ticket Guy"
    assert display_name("YS Needle Tickets") == "Needle"
    assert display_name("YSW Tickets") == "YSW (Waxler)"


def test_display_name_falls_back_for_unknown_qbo_name():
    """If a new QBO company shows up in the master before the dict is updated,
    we fall back to the QBO name itself rather than dropping the row."""
    assert display_name("Some Brand New Co") == "Some Brand New Co"


def test_summary_orders_companies_with_yourtickets_last():
    """Summary order must match the Purchase Details processor: Y&S, Grossman,
    Sternbuch, Pollak, ..., Damona, YourTickets (always last)."""
    df = pd.DataFrame([
        _raw_row(Company="YourTickets", **{"PO #": 1, "Total Adjustment":  5.0}),
        _raw_row(Company="YSA",         **{"PO #": 2, "Total Adjustment": 10.0}),
        _raw_row(Company="YS Tickets",  **{"PO #": 3, "Total Adjustment": 20.0}),
        _raw_row(Company="Damon and Crew", **{"PO #": 4, "Total Adjustment": 30.0}),
    ])
    # Extend the test mapping to include all four
    mapping = {**TEST_MAPPING,
               "yourtickets": "YourTickets",
               "damon and crew": "Damona & Crew"}
    out, _ = transform(df, mapping=mapping)
    s = summarize_by_company(out)
    order = [c["company"] for c in s["companies"]]
    # YourTickets should be last; Y&S should come before YSA; Damona before YourTickets.
    assert order[-1] == "YourTickets"
    assert order.index("Y&S") < order.index("YSA")
    assert order.index("Damona") < order.index("YourTickets")


def test_real_master_mapping_file_loads():
    """The real Master_Mapping_List.xlsx should load with expected mappings.
    Keys are lowercased; values keep canonical casing."""
    from mapping import load_mapping
    m = load_mapping("data/Master_Mapping_List.xlsx")
    assert m.get("ysa") == "YS Asher Tickets"
    assert m.get("ysa 2") == "YS Asher Tickets"
    assert m.get("jacks ys") == "YS Chase Tickets"
    assert m.get("ys tickets") == "Y&S Tickets"
    # The real-world case bug: master stored "YS-SeatGeek2", data uses "YS-Seatgeek2".
    # After lowercasing, both forms now resolve to the same key.
    assert m.get("ys-seatgeek2") == "Y&S Tickets"


# ---------------------------------------------------------------------------
# Bills + Expenses ledger
# ---------------------------------------------------------------------------

def test_bills_and_expenses_split_by_sign():
    """Positive adjustments go to Bills (1 row each).
    Negative adjustments go to Expenses (2 rows each, summing to zero)."""
    from processor import _build_bills_and_expenses

    df = pd.DataFrame([
        _raw_row(Company="YSA",        **{"PO #": 100, "Total Adjustment": 200.0, "Vendor": "Ticketmaster"}),
        _raw_row(Company="YS Tickets", **{"PO #": 101, "Total Adjustment": -150.0, "Vendor": "Ticketmaster"}),
    ])
    cleaned, _ = transform(df, mapping=TEST_MAPPING)
    bills, expenses = _build_bills_and_expenses(cleaned, open_vendors=set())

    assert len(bills) == 1
    assert bills.iloc[0]["Total"] == 200.0
    assert bills.iloc[0]["Category"] == "Inventory Asset"
    assert bills.iloc[0]["Account"] == "Clearing Account"

    assert len(expenses) == 2
    assert expenses.iloc[0]["Total"] + expenses.iloc[1]["Total"] == 0.0
    assert expenses.iloc[0]["Category"] == "Inventory Asset"
    assert expenses.iloc[0]["Total"] == -150.0
    assert expenses.iloc[1]["Total"] == 150.0


def test_expenses_offset_category_uses_open_vendors_set():
    """Vendors in the open-vendors set → 'Due from Vendors - Open'.
    Other vendors → '<Vendor> (TC)'. Case-insensitive."""
    from processor import _build_bills_and_expenses

    df = pd.DataFrame([
        _raw_row(Company="YSA", **{"PO #": 1, "Total Adjustment": -100.0, "Vendor": "Ticketmaster"}),
        _raw_row(Company="YSA", **{"PO #": 2, "Total Adjustment": -50.0,  "Vendor": "Concert Extras"}),
        _raw_row(Company="YSA", **{"PO #": 3, "Total Adjustment": -25.0,  "Vendor": "TICKETMASTER"}),  # case test
    ])
    cleaned, _ = transform(df, mapping=TEST_MAPPING)
    bills, expenses = _build_bills_and_expenses(
        cleaned, open_vendors={"ticketmaster"}  # lowercased
    )

    # Find the offset-line (line 2) for each PO and check its Category.
    offset_lines = expenses[expenses["Category"] != "Inventory Asset"]
    by_po = dict(zip(offset_lines["Expense #"], offset_lines["Category"]))

    tm_categories = [v for k, v in by_po.items()
                     if expenses[expenses["Expense #"] == k].iloc[0]["Vendor"].lower() == "ticketmaster"]
    assert all(c == "Due from Vendors - Open" for c in tm_categories)

    ce_categories = [v for k, v in by_po.items()
                     if expenses[expenses["Expense #"] == k].iloc[0]["Vendor"] == "Concert Extras"]
    assert ce_categories == ["Concert Extras (TC)"]


def test_expense_numbering_is_global_and_sequential():
    """Expense # is a single global counter across Bills + Expenses."""
    from processor import _build_bills_and_expenses

    df = pd.DataFrame([
        _raw_row(Company="YSA",        **{"PO #": 100, "Total Adjustment":  10.0, "Team/Performer": "A"}),
        _raw_row(Company="YS Tickets", **{"PO #": 200, "Total Adjustment": -20.0, "Team/Performer": "B"}),
        _raw_row(Company="YSA",        **{"PO #": 300, "Total Adjustment": -30.0, "Team/Performer": "C"}),
    ])
    cleaned, _ = transform(df, mapping=TEST_MAPPING)
    bills, expenses = _build_bills_and_expenses(cleaned, open_vendors=set())

    all_expense_nums = sorted(set(bills["Expense #"]).union(expenses["Expense #"]))
    assert all_expense_nums == [1, 2, 3]


def test_company_column_uses_original_ticketvault_name():
    """Bills/Expenses 'Company' column = the ORIGINAL company name from the
    upload, not the display label or QBO name."""
    from processor import _build_bills_and_expenses

    df = pd.DataFrame([
        _raw_row(Company="YSA 2", **{"PO #": 100, "Total Adjustment": -50.0}),
    ])
    cleaned, _ = transform(df, mapping=TEST_MAPPING)
    bills, expenses = _build_bills_and_expenses(cleaned, open_vendors=set())

    assert set(expenses["Company"].unique()) == {"YSA 2"}


def test_memo_format():
    """Memo/Description = '<Team/Performer> - Cost Changes (<Original Company>)'"""
    from processor import _build_bills_and_expenses

    df = pd.DataFrame([
        _raw_row(Company="YS-Seatgeek2", **{
            "PO #": 100, "Total Adjustment": -50.0,
            "Team/Performer": "New York Knicks",
        }),
    ])
    cleaned, _ = transform(df, mapping=TEST_MAPPING)
    _, expenses = _build_bills_and_expenses(cleaned, open_vendors=set())
    expected = "New York Knicks - Cost Changes (YS-Seatgeek2)"
    assert expenses.iloc[0]["Memo"] == expected
    assert expenses.iloc[0]["Description"] == expected


def test_zero_rows_dont_appear_anywhere():
    """Zero adjustments are filtered out in transform; they should never
    reach Bills or Expenses."""
    from processor import _build_bills_and_expenses

    df = pd.DataFrame([
        _raw_row(Company="YSA", **{"PO #": 100, "Total Adjustment": 0.0,
                                    "Ticket Cost Total End": 100.0}),
        _raw_row(Company="YSA", **{"PO #": 101, "Total Adjustment": 10.0}),
    ])
    cleaned, _ = transform(df, mapping=TEST_MAPPING)
    bills, expenses = _build_bills_and_expenses(cleaned, open_vendors=set())
    assert len(bills) == 1
    assert len(expenses) == 0


# ---------------------------------------------------------------------------
# Ticketmaster AM → team-performer rename
# ---------------------------------------------------------------------------

def test_ticketmaster_am_rule_renames_when_performer_is_a_team():
    """Vendor 'Ticketmaster AM' + known team performer → Vendor becomes the team."""
    from teams import rename_ticketmaster_am
    teams = {"atlanta hawks", "cleveland cavaliers"}
    assert rename_ticketmaster_am("Ticketmaster AM", "Atlanta Hawks", teams) == "Atlanta Hawks"
    assert rename_ticketmaster_am("ticketmaster am", "atlanta hawks", teams) == "atlanta hawks"
    assert rename_ticketmaster_am("Ticketmaster AM", "CLEVELAND CAVALIERS", teams) == "CLEVELAND CAVALIERS"


def test_ticketmaster_am_rule_leaves_vendor_when_performer_is_not_a_team():
    """Performer that's not in the team list → Vendor stays 'Ticketmaster AM'."""
    from teams import rename_ticketmaster_am
    teams = {"atlanta hawks"}
    assert rename_ticketmaster_am("Ticketmaster AM", "Rosalía", teams) == "Ticketmaster AM"
    assert rename_ticketmaster_am("Ticketmaster AM", None, teams) == "Ticketmaster AM"


def test_ticketmaster_am_rule_leaves_other_vendors_alone():
    """Vendors that aren't 'Ticketmaster AM' are untouched even if performer is a team."""
    from teams import rename_ticketmaster_am
    teams = {"atlanta hawks"}
    assert rename_ticketmaster_am("Ticketmaster",   "Atlanta Hawks", teams) == "Ticketmaster"
    assert rename_ticketmaster_am("Live Nation",    "Atlanta Hawks", teams) == "Live Nation"
    assert rename_ticketmaster_am("Concert Extras", "Atlanta Hawks", teams) == "Concert Extras"


def test_ticketmaster_am_rule_drives_aggregation_in_transform():
    """End-to-end: two rows that only differ in Vendor (Ticketmaster AM vs the
    matching team name) should aggregate into one row after the rename."""
    df = pd.DataFrame([
        _raw_row(Company="YSA", **{
            "PO #": 1, "Total Adjustment": 50.0,
            "Vendor": "Ticketmaster AM", "Team/Performer": "Atlanta Hawks",
        }),
        _raw_row(Company="YSA", **{
            "PO #": 2, "Total Adjustment": 30.0,
            "Vendor": "Atlanta Hawks",   "Team/Performer": "Atlanta Hawks",
        }),
    ])
    # We need teams to include Atlanta Hawks; can't pass to transform() so
    # monkey-patch the cached getter.
    import teams as teams_mod
    teams_mod.reset_cache()
    original = teams_mod.get_teams
    teams_mod.get_teams = lambda: {"atlanta hawks"}
    try:
        out, _ = transform(df, mapping=TEST_MAPPING)
    finally:
        teams_mod.get_teams = original
        teams_mod.reset_cache()
    assert len(out) == 1
    assert out.iloc[0]["Vendor"] == "Atlanta Hawks"
    assert out.iloc[0]["Total Adjustment"] == 80.0


def test_team_rename_rule_covers_tickets_dot_com_and_ballpark():
    """Tickets.com and Ballpark follow the same team-rename rule as Ticketmaster AM."""
    from teams import rename_team_vendor
    teams = {"atlanta hawks", "new york yankees"}
    assert rename_team_vendor("Tickets.com",   "Atlanta Hawks",     teams) == "Atlanta Hawks"
    assert rename_team_vendor("tickets.com",   "New York Yankees",  teams) == "New York Yankees"
    assert rename_team_vendor("Ballpark",      "Atlanta Hawks",     teams) == "Atlanta Hawks"
    assert rename_team_vendor("BALLPARK",      "New York Yankees",  teams) == "New York Yankees"
    # Non-team performer → stays unchanged
    assert rename_team_vendor("Tickets.com",   "Some Concert",      teams) == "Tickets.com"
    assert rename_team_vendor("Ballpark",      "Some Concert",      teams) == "Ballpark"


def test_real_teams_file_loads_with_expected_leagues():
    """Sanity check: the real major_league_teams.xlsx should have ~150 teams
    spanning NBA / NFL / NHL / MLB / MLS."""
    from teams import load_teams
    t = load_teams("data/major_league_teams.xlsx")
    assert "atlanta hawks" in t
    assert "new york yankees" in t
    assert "dallas cowboys" in t
    assert 100 < len(t) < 200


# ---------------------------------------------------------------------------
# Purchase Details exclusion
# ---------------------------------------------------------------------------

def test_process_files_excludes_pos_present_in_purchase_details():
    """End-to-end: when a PD file lists certain PO #s, those rows must be
    excluded from output (Combined / Bills / per-company tabs)."""
    import io
    import openpyxl
    from processor import process_files

    cc_rows = pd.DataFrame([
        _raw_row(Company="YSA",        **{"PO #": 1001, "Total Adjustment": 50.0,  "Team/Performer": "A"}),
        _raw_row(Company="YS Tickets", **{"PO #": 1002, "Total Adjustment": -75.0, "Team/Performer": "B"}),
        _raw_row(Company="YSA 2",      **{"PO #": 1003, "Total Adjustment": 25.0,  "Team/Performer": "C"}),
    ])
    cc_buf = io.BytesIO()
    cc_rows.to_csv(cc_buf, index=False)
    cc_buf.seek(0)

    pd_rows = pd.DataFrame({
        "Company": ["YSA"],
        "PO #":   [1001],   # Only PO 1001 — should remove just that row from output
    })
    pd_buf = io.BytesIO()
    pd_rows.to_csv(pd_buf, index=False)
    pd_buf.seek(0)

    # CC files have a two-row header in practice; our pandas DataFrame goes
    # straight to CSV with single-row headers, so we use that path directly.
    # The reader auto-detects header=0 vs header=1.
    result = process_files(
        [(cc_buf.getvalue(), "cc.csv")],
        pd_file_list=[(pd_buf.getvalue(), "pd.csv")],
    )

    assert result["excluded"]["row_count"] == 1
    assert result["excluded"]["po_count"] == 1

    # Combined ledger should have 2 events (1002, 1003), not 3.
    wb = openpyxl.load_workbook(io.BytesIO(result["combined"]))
    combined = wb["Combined"]
    assert combined.max_row - 1 == 2  # minus header

    # Excluded tab should exist with 1 row.
    assert "Excluded" in wb.sheetnames
    assert wb["Excluded"].max_row - 1 == 1
